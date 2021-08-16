import os
import sys
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib
import shutil
import configparser
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series, BarChart

#path = os.path.dirname(os.path.abspath(__file__))
SD = "my_calculation_module_directory"
path = os.path.dirname(os.path.abspath(__file__)).split(SD)[0] + "/%s" % SD

if path not in sys.path:
    sys.path.append(path)
# from CM.helper_functions.readCsvData import READ_CSV_DATA

def linear_interpol(y0,x0, y1,x1, x):
    return (y0*(x1-x) + y1*(x-x0))/(x1-x0)

def dfToCSV(results, output_dir, scen, filename):
    output_dir = os.path.join(output_dir,scen + '_final')
    if os.path.exists(output_dir):
        shutil.rmtree(output_dir)
        os.mkdir(output_dir)
    else:
        os.mkdir(output_dir)
    output_file = os.path.join(output_dir,filename)
    results.transpose().to_csv(output_file)
    return output_dir

def printPlots(results, output_directory):
    """
    Estimated heated area total: gfa_fut
    Estimated heated area per construction period: gfa_xx_fut
    Estimated heated area per capita: gfa_per_cap_fut

    Energy Consumption total: ene_fut
    Energy Consumption per construction period: ene_xx_fut

    Estimated specific Energy Consumption: spec_ene_fut
    Estimated specific Energy Consumption per construction period: spec_ene_xx_fut

    Estimated population grow: pop_fut
    """

    ##########
    # Heated area plot
    ##########
    matplotlib.rcParams['interactive'] == False
    results = results.transpose()
    year = list(results.index)

    fig = plt.figure(figsize = (10,5))
    ax1 = plt.subplot(1,2,1)
    ax1.plot(year, results['gfa_fut'] / 10 ** 6, label='total', color='blue')
    ax1.set(xlabel='year', ylabel='total estimated heated area [Mio m2]')
    ax1.grid()

    ax2 = plt.subplot(1, 2, 2)
    ax2.plot(year, results['gfa_75_fut']/10**6, label='75', color = 'red')
    ax2.plot(year, results['gfa_80_fut']/10**6, label='80', color =  'orange')
    ax2.plot(year, results['gfa_00_fut']/10**6, label='00', color =  'green')
    ax2.plot(year, results['gfa_new_fut']/10**6, label='new_buildings', color =  'purple')
    ax2.legend()
    ax2.set(xlabel='year', ylabel='Estimated heated area per construction period [Mio m2]')
    ax2.grid()
    plt.tight_layout()
    fig.savefig(os.path.join(output_directory,"Estimated_heated_area.png"))
    plt.close()
    #plt.show()

    ##########
    # Heated are per capita plot
    ##########
    fig, ax = plt.subplots()
    ax.plot(year, results['gfa_per_cap_fut'], label='gfa_per_cap')
    ax.set(xlabel='year', ylabel='Estimated heated area per capita [m2/capita]')
    ax.grid()
    fig.savefig(os.path.join(output_directory, "Estimated_specific_energy_consumption.png"))
    plt.close()
    #plt.show()

    ##########
    # Energy consumption plot
    ##########
    fig = plt.figure(figsize = (10,5))
    ax1 = plt.subplot(1,2,1)
    ax1.plot(year, results['ene_fut'] / 10 ** 3, label='total')
    ax1.set(xlabel='year', ylabel='Total estimated Energy Consumption [GWh]')
    ax1.grid()

    ax2 = plt.subplot(1, 2, 2)
    ax2.plot(year, results['ene_75_fut'] / 10 ** 3, label='75')
    ax2.plot(year, results['ene_80_fut'] / 10 ** 3, label='80')
    ax2.plot(year, results['ene_00_fut'] / 10 ** 3, label='00')
    ax2.plot(year, results['ene_new_fut'] / 10 ** 3, label='new_buildings')
    ax2.set(xlabel='year', ylabel='Estimated Energy Consumption per construction period [GWh]')
    ax2.grid()
    ax2.legend()
    fig.savefig(os.path.join(output_directory, "Estimated_energy_consumption.png"))
    plt.close()
    #plt.show()

    ##########
    # Specific Energy consumption plot
    ##########

    fig, ax = plt.subplots()
    ax.plot(year, results['spe_ene_fut'], label='total')
    ax.plot(year, results['spec_ene_75_fut'], label='75')
    ax.plot(year, results['spec_ene_80_fut'], label='80')
    ax.plot(year, results['spec_ene_00_fut'], label='00')
    ax.plot(year, results['spec_ene_new_fut'], label='new')
    ax.set(xlabel='year', ylabel='Specific Energy Consumption [kWh/m2]')
    ax.grid()
    plt.legend()
    fig.savefig(os.path.join(output_directory, "Estimated_specific_energy_consumption.png"))
    plt.close()
    #plt.show()

    ##########
    # Human population grow plot
    ##########

    fig, ax = plt.subplots()
    ax.plot(year, results['pop_fut'], label='population')
    ax.set(xlabel='year', ylabel='Estimated population grow [tsd. people]')
    ax.grid()
    fig.savefig(os.path.join(output_directory, "Estimated_population_grow.png"))
    plt.close()
    #plt.show()
    print("Exported plots to '%s'" % output_directory)

def createLINECHART(sheet, sheetname, title, style, x_axis_title, y_axis_title, labels, xvalues, yvalues):
    chart = ScatterChart()
    chart.title = title
    chart.style = style
    chart.x_axis.title = x_axis_title
    chart.y_axis.title = y_axis_title
    # chart.x_axis.scaling.min = 
    # chart.x_axis.scaling.max = 
    xvalues = Reference(sheet, 
                        min_col=xvalues['min_col'],
                        min_row=xvalues['min_row'], 
                        max_col=xvalues['max_col'])
    if labels is None: 
        labels = [None]*len(yvalues['min_row'])
        chart.legend = None
    for i, label in zip(yvalues['min_row'], labels): 
        values = Reference(sheet, 
                           min_col=yvalues['min_col'], 
                           min_row = i, 
                           max_col = yvalues['max_col'])
        series = Series(values, xvalues, title_from_data = False, 
                        title = label)
        series.graphicalProperties.line.width = 30000
        chart.series.append(series)
    # if labels[0] is None: chart.legend = None
    return chart

def createBARCHART(sheet, sheetname, title, style, x_axis_title, y_axis_title, labels, xvalues, yvalues): 
    chart = BarChart()
    chart.type = 'col'
    chart.title = title
    chart.style = style
    if 'GFA' in title or 'Pop' in title: 
        chart.legend = None
    chart.x_axis.title = x_axis_title
    chart.y_axis.title = y_axis_title
    xvalues = Reference(sheet, 
                        min_col=xvalues['min_col'],
                        min_row=xvalues['min_row'], 
                        max_col=xvalues['max_col'], 
                        max_row=xvalues['max_row'])
    for col, label in zip(yvalues['min_col'], labels): 
        values = Reference(sheet, 
                           min_col = col,
                           min_row = yvalues['min_row'], 
                           max_col = col,
                           max_row = yvalues['max_row'])
        series = Series(values, title = label)
        chart.append(series)
    chart.set_categories(xvalues)
    chart.x_axis.tickLblPos ='high' #'nextTO', 'high', 'low'
    chart.shape = 4 
    return chart
    

def createASSESCHART(sheet, sheetname, title, style, x_axis_title, y_axis_title, labels, xvalues, yvalues):
    chart = ScatterChart()
    chart.title = title
    chart.style = style
    chart.x_axis.title = x_axis_title
    chart.y_axis.title = y_axis_title
    # chart.x_axis.scaling.min = 
    # chart.x_axis.scaling.max = 
    xvalues = Reference(sheet, 
                        min_col=xvalues['min_col'],
                        min_row=xvalues['min_row'], 
                        max_col=xvalues['max_col'])
    if labels is None: 
        labels = [None]*len(yvalues['min_row'])
        chart.legend = None
    for i, label in zip(yvalues['min_col'], labels): 
        values = Reference(sheet, 
                           min_col=i, 
                           min_row = yvalues['min_row'], 
                           max_col = i+1)
        series = Series(values, xvalues, title_from_data = False, 
                        title = label)
        series.graphicalProperties.line.width = 40000
        chart.series.append(series)
    return chart
    
def create_CHARTS(file,sheet_name, results):
    wb = openpyxl.load_workbook(file)
    scen_list = []
    for col in results.columns: 
        if not col[5:] in scen_list: scen_list.append(col[5:])
    try: 
        xvalues = {'min_col': 2, 'min_row':49, 'max_col':3}
        
        # GFA in total, res, nres
        yvalues = {'min_col': 2, 'min_row':range(2,5), 'max_col':3}
        labels = results.index[0:3]
        chart = createLINECHART(wb[sheet_name], sheet_name, '', 13, 'year', 
                            'gross floor area [m2]', labels, xvalues, yvalues)
        wb[sheet_name].add_chart(chart, "AC1")
        
        # GFA per cp in total, res, nres
        yvalues = {'min_col': 2, 'min_row':range(11,23), 'max_col':3}
        labels = results.index[9:21]
        chart = createLINECHART(wb[sheet_name], sheet_name, '', 13, 'year', 
                            'gross floor area [m2]', labels, xvalues, yvalues)
        wb[sheet_name].add_chart(chart, "AL1")
        
        # GFA per capita total
        yvalues = {'min_col': 2, 'min_row':[48], 'max_col':3}
        labels = None
        chart = createLINECHART(wb[sheet_name], sheet_name, '', 13, 'year', 
                            'gross floor area/capita', labels, xvalues, yvalues)
        wb[sheet_name].add_chart(chart, "AU1")
        
        # EnergyDemand in total, res and nres
        chart_positions = ['AC16','AL16','AU16','BD16','BM16', 'BV16']
        for chart_pos, col_i, scen in zip(chart_positions, range(2,14,2), scen_list):
            yvalues = {'min_col': col_i, 'min_row':range(5,8), 'max_col':(col_i+1)}
            labels = results.index[3:6]
            chart = createLINECHART(wb[sheet_name], sheet_name, scen, 13, 'year', 
                                'energy demand [kWh]', labels, xvalues, yvalues)
            wb[sheet_name].add_chart(chart, chart_pos) 
        
        # EnergyDemand in total all scenarios
        yvalues = {'min_col': range(2,14,2), 'min_row':5, 'max_col':None}
        labels = scen_list
        chart = createASSESCHART(wb[sheet_name], sheet_name, 'energy demand_total', 13, 'year', 
                                  'energy demand [kWh]', labels, xvalues, yvalues)
        wb[sheet_name].add_chart(chart, 'CE16')
            
        # EnergyDemand per cp in total, res and nres 
        chart_positions = ['AC31','AL31','AU31','BD31','BM31', 'BV31']
        for chart_pos, col_i, scen in zip(chart_positions, range(2,14,2), scen_list):
            yvalues = {'min_col': col_i, 'min_row':range(23,35), 'max_col':(col_i+1)}
            labels = results.index[21:33]
            chart = createLINECHART(wb[sheet_name], sheet_name, scen, 13, 'year', 
                                'energy demand [kWh]', labels, xvalues, yvalues)
            wb[sheet_name].add_chart(chart, chart_pos) 
            
        # Specific EnergyDemand in total, res and nres
        chart_positions = ['AC46','AL46','AU46','BD46','BM46', 'BV46']
        for chart_pos, col_i, scen in zip(chart_positions, range(2,14,2), scen_list):
            yvalues = {'min_col': col_i, 'min_row':range(8,11), 'max_col':(col_i+1)}
            labels = results.index[6:9]
            chart = createLINECHART(wb[sheet_name], sheet_name, scen, 13, 'year', 
                                'energy demand/m2 [kWh/m2]', labels, xvalues, yvalues)
            wb[sheet_name].add_chart(chart, chart_pos) 
            
        # Specific EnergyDemand in total all scenarios
        yvalues = {'min_col': range(2,14,2), 'min_row':8, 'max_col':None}
        labels = scen_list
        chart = createASSESCHART(wb[sheet_name], sheet_name, 'specific energy demand_total', 13, 'year', 
                                  'specific energy demand [kWh/m2]', labels, xvalues, yvalues)
        wb[sheet_name].add_chart(chart, 'CE46')
        
        # Specific EnergyDemand per cp in total, res and nres
        chart_positions =  ['AC61','AL61','AU61','BD61','BM61', 'BV61']
        for chart_pos, col_i, scen in zip(chart_positions, range(2,14,2), scen_list):
            yvalues = {'min_col': col_i, 'min_row':range(35,47), 'max_col':(col_i+1)}
            labels = results.index[33:45]
            chart = createLINECHART(wb[sheet_name], sheet_name, scen, 13, 'year', 
                                'energy demand/m2 [kWh/m2]', labels, xvalues, yvalues)
            wb[sheet_name].add_chart(chart, chart_pos)
        
        # GFA in total,res and nres, relativ 
        chart_pos = ['T1']
        xvalues = {'min_col': 1, 'min_row':2, 'max_col':1, 'max_row':4}
        yvalues = {'min_col': [14], 'min_row':2, 'max_col':14, 'max_row':4}
        chart = createBARCHART(wb[sheet_name], sheet_name, 'Relativ GFA', 10, 'Categories', 'Rel. Change [%]', scen_list, xvalues, yvalues)
        wb[sheet_name].add_chart(chart, chart_pos[0])
        
        # ENE in total,res and nres, relativ 
        chart_pos = ['T16']
        xvalues = {'min_col': 1, 'min_row':5, 'max_col':1, 'max_row':7}
        yvalues = {'min_col': range(14,20), 'min_row':5, 'max_col':14, 'max_row':7}
        chart = createBARCHART(wb[sheet_name], sheet_name, 'Relativ ENE', 10, 'Categories', 'Rel. Change [%]', scen_list, xvalues, yvalues)
        wb[sheet_name].add_chart(chart, chart_pos[0])
        
        # SPE_ENE in total,res and nres, relativ 
        chart_pos = ['T31']
        xvalues = {'min_col': 1, 'min_row':8, 'max_col':1, 'max_row':10}
        yvalues = {'min_col': range(14,20), 'min_row':8, 'max_col':14, 'max_row':10}
        chart = createBARCHART(wb[sheet_name], sheet_name, 'Relativ SPE_ENE', 10, 'Categories', 'Rel. Change [%]', scen_list, xvalues, yvalues)
        wb[sheet_name].add_chart(chart, chart_pos[0])
        
        # POP, relativ 
        chart_pos = ['T46']
        xvalues = {'min_col': 1, 'min_row':47, 'max_col':1, 'max_row':47}
        yvalues = {'min_col': [14], 'min_row':47, 'max_col':14, 'max_row':47}
        chart = createBARCHART(wb[sheet_name], sheet_name, 'Relativ Population Growth', 10, '', 'Rel. Change [%]', scen_list, xvalues, yvalues)
        wb[sheet_name].add_chart(chart, chart_pos[0])
        
        wb.save(file)
    except Exception as e: 
        wb.close()
        raise e
    wb.close()

def dataToEXCEL(scen_list, output_dir, network, input_dir):
    filename = network + '_results.xlsx'
    # if not path.exists(os.path.join(out_dir,filename)):
    #     xlsxfile = openpyxl.Workbook()
    results = pd.DataFrame(data = {})
    for scen in scen_list.keys():
        # scen_list[scen].sort()
        if len(scen_list[scen]) != 2: continue
        input_scen_path_1 = os.path.join(input_dir,
                                         scen +'_'+ scen_list[scen][0],
                                         'indicators_exact.csv')
        input_scen_path_2 = os.path.join(input_dir,
                                         scen + '_' + scen_list[scen][1], 
                                         'indicators_exact.csv')
        results_baseyear = pd.read_csv(input_scen_path_1,
                                       index_col = 0, 
                                       header = 0, 
                                       skiprows = [1]) 
        results_futureyear = pd.read_csv(input_scen_path_2,
                                         index_col = 0, 
                                         header = 0, 
                                         skiprows = [1])
        baseyear = (str(int(results_baseyear.loc['target_year'][0])) +
                    "_" + scen[39:])
        futureyear = (str(int(results_futureyear.loc['target_year'][0])) + 
                     "_" + scen[39:])
        results_baseyear.rename(columns={results_baseyear.columns[0]:baseyear},
                                inplace = True)
        results_futureyear.rename(columns={results_futureyear.columns[0]:futureyear},
                                  inplace = True)
        if len(results.index) == 0: 
            # Assume that all values from baseyear are basically the same in each scenario 
            # (they are not the same, deviation of 0.002% from each other)
            results = results_baseyear.copy()
        else: 
            results = results.join(results_baseyear)
        results = results.join(results_futureyear)
    dropindex_lst = []
    for index in results.index: 
        if 'cur' in index: dropindex_lst.append(index)
    results.drop(index = dropindex_lst, inplace = True)
    results.set_index(results.index.str.replace('_fut', ''), inplace = True)
    dropindex_lst = config['drop_from_results']['drop_index'].split('\n')[1:]
    results.drop(index = dropindex_lst, inplace = True)
    results.index.name = ''
    #results.reset_index(inplace = True, drop = False)
    for key in scen_list.keys():
        try: 
            results[key[39:]+'_rel'] = results['2050_'+key[39:]]/results['2017_'+key[39:]]-1
            results[key[39:]+'_rel'] = round(results[key[39:]+'_rel']*100,2)
        except KeyError: 
            del scen_list[key]
            
    
    filename = network + '_results.xlsx'
    file = os.path.join(output_dir, filename)
    sheet_name = os.path.split(input_dir)[1][5:] + '_reg_data'
    if len(sheet_name)>= 31:
        sheet_name = os.path.split(input_dir)[1][5:26] + '_reg_data'
    file_path = os.path.join(output_dir,filename)
    if not os.path.exists(file_path):
        mode = "w"
        with pd.ExcelWriter(file, 
                    engine = 'openpyxl', 
                    mode = mode) as writer:
            results.to_excel(writer, 
                             sheet_name= sheet_name,
                             index = True)
    else:
        workbook= openpyxl.load_workbook(file)
        if sheet_name in workbook.sheetnames and len(workbook.sheetnames)>1: 
            del workbook[sheet_name]
            with pd.ExcelWriter(file, 
                            engine = 'openpyxl', 
                            mode = 'a', 
                            if_sheet_exists = 'replace') as writer:
                results.to_excel(writer,
                                 sheet_name= sheet_name,
                                 index = True)
        elif sheet_name in workbook.sheetnames and len(workbook.sheetnames)==1:
            os.remove(file_path)
            with pd.ExcelWriter(file,
                                engine = 'openpyxl',
                                mode = 'w') as writer:
                results.to_excel(writer,
                                 sheet_name= sheet_name,
                                 index = True)
        elif not sheet_name in workbook.sheetnames: 
             with pd.ExcelWriter(file,
                                engine = 'openpyxl',
                                mode = 'a',
                                if_sheet_exists = 'error') as writer:
                results.to_excel(writer,
                                 sheet_name= sheet_name,
                                 index = True)
        else:
            print("Error: Not able to create sheet %s, continue with next"
                  % sheet_name)
            return None
    create_CHARTS(file,sheet_name, results)
    return results

    
def main(network, input_directory, output_dir, interpol):
    regions = os.listdir(input_directory)
    for region in regions:
        if not (network in region):
            continue
        input_directory_ = os.path.join(input_directory, region)
        dir_list = os.listdir(input_directory_)
        dir_list.sort()
        scen_list = {}
        year_list = []
        for item in dir_list:
            if '.' in item: continue
            if item[:-5] not in scen_list.keys():
                scen_list[item[:-5]] = []
            if item[-4:] not in scen_list[item[:-5]]:
                scen_list[item[:-5]].append(item[-4:])
        dataToEXCEL(scen_list,
                    output_dir, 
                    network,
                    input_directory_)
        #break #DEBUG break
            # if interpol == 'linear':
            #     for year in range(2020, 2050,5):
            #         results[year] = linear_interpol(results_baseyear[baseyear],baseyear,
            #                                         results[futureyear],futureyear,
            #                                         year)
            # else:
            #     print(interpol + " is not implemented. Use 'linear' instead")
            # out_path_final = dfToCSV(results, output_dir,scen, 'results.csv')
            # printPlots(results,out_path_final)
        
def modProjectPath(path):
    proj_path = os.getcwd().split('projects2')
    return os.path.join(proj_path[0],path)

if __name__ == '__main__':
    configfile_location = 'result_preperation_conf.ini'
    config = configparser.ConfigParser()
    if not config.read(configfile_location):
        raise AttributeError("Couldn't load configuration file ",
                             configfile_location)
    if config['machine']['machine'] == 'server':
        input_paths_label = 'proj_path_' + 'server'
        proj_path = modProjectPath(config[input_paths_label]['proj_path'])
    elif config['machine']['machine'] == 'local':
        input_paths_label = 'proj_path_' + 'local'
        proj_path = str(config[input_paths_label]['proj_path'])
    else:
        raise AttributeError("No selected machine in configuration file")
    data_warehouse = os.path.join(proj_path, config['input_paths']['data_warehouse'])
    output_dir = os.path.join(proj_path, config['input_paths']['output_dir'])
    if not os.path.exists(data_warehouse):
        raise FileNotFoundError(data_warehouse)
    regions = os.listdir(data_warehouse)
    networks = list(set([network[:2] for network in regions]))
    networks_to_prep = config['input_var']['networks_to_prep'].split('\n')[1:]
    for network in networks:
        if not network in networks_to_prep: continue
        print("Start to prepare Results for network in %s" % network)
        # directory = os.path.join(data_warehouse, region)
        #input_dir_ = "E:/projects2/bmayr_workspace/projects_bernhard/outputs/res_h_regions_test/FR - CCSB"
        output_dir_ = os.path.join(output_dir,network)
        if not os.path.exists(output_dir_):
            try:
                os.mkdir(output_dir_)
            except FileNotFoundError:
                os.makedirs(output_dir_)
        interpol_ = 'linear'
        main(network, data_warehouse, output_dir_, interpol_)
        print("Finished with result-preparation for network in %s" % network)
        print('-'*10)
    print("DONE")
    print("#"*10)